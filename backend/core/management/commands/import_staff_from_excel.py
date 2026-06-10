"""
usage: python manage.py import_staff_from_excel <excel_path>
"""
import sys
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from core.models import Shop
from core.models.base import User

ROLE_MAP = {
    "①": "executive",
    "②": "accounting",
    "③": "manager",
    "④": "store_manager",
    "⑤": "staff",
}

HEADER_SENTINEL = "権限設定"


class Command(BaseCommand):
    help = "Excelファイルからスタッフを一括インポートする"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Excelファイルのパス")

    def handle(self, *args, **options):
        path = options["excel_path"]

        # ── 店舗名 → Shop オブジェクトのキャッシュ ──
        shop_cache = {s.name: s for s in Shop.objects.all()}

        wb = load_workbook(path)
        ws = wb.active

        created = []
        skipped = []
        unknown_shops = set()

        for row in ws.iter_rows(values_only=True):
            shop_name = row[1]
            code      = row[2]
            name      = row[3]
            role_raw  = row[4]

            # ヘッダー行・空行をスキップ
            if not code or not name or not role_raw:
                continue
            role_raw = str(role_raw).strip()
            if HEADER_SENTINEL in role_raw:
                continue

            login_id     = str(code).strip()
            display_name = str(name).strip()
            role         = ROLE_MAP.get(role_raw, "staff")
            shop_str     = str(shop_name).strip() if shop_name else ""

            # 店舗マッチング
            shop = shop_cache.get(shop_str)
            if shop_str and not shop:
                unknown_shops.add(shop_str)

            # 既存ログインIDはスキップ
            if User.objects.filter(login_id=login_id).exists():
                skipped.append(login_id)
                continue

            user = User(
                login_id=login_id,
                display_name=display_name,
                role=role,
                shop=shop,
                is_staff=True,
            )
            user.set_password(login_id)  # 初期パスワード = 従業員コード
            user.save()
            created.append((login_id, display_name, role, shop_str))

        # ── 結果レポート ──
        self.stdout.write(self.style.SUCCESS(f"\n✅ 作成: {len(created)}件"))
        for login_id, name, role, shop in created:
            self.stdout.write(f"  {login_id:>5}  {name:<20}  {role:<14}  {shop}")

        if skipped:
            self.stdout.write(self.style.WARNING(f"\n⏭️  スキップ（既存）: {len(skipped)}件"))
            self.stdout.write("  " + ", ".join(skipped))

        if unknown_shops:
            self.stdout.write(self.style.WARNING(f"\n⚠️  DBに未登録の店舗（shop=未設定でインポート済み）:"))
            for s in sorted(unknown_shops):
                self.stdout.write(f"  {s}")

        self.stdout.write(self.style.SUCCESS("\n初期パスワード = 従業員コード（各自変更を推奨）\n"))
